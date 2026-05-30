"""
Extract question fingerprints from:
A. Chapter_Tests/Tests/ — .docx files
B. Practice_Tests_by_Topic/Kristen_Topic_Tests/ — files WITHOUT "answer"/"explanation" in filename
C. Practice_Exams/Kristen_Mini_Exams/ — files WITHOUT "answer"/"explanation" in filename
D. Practice_Exams/Past_ABT_Exams/ — .docx and .xlsx files (skip .pptx and answer-only)

Fingerprint: first 80 chars after cleaning (lowercase, strip punctuation, collapse whitespace)
"""

import os
import re
import csv
import glob
import sys

try:
    from docx import Document
except ImportError:
    print("ERROR: python-docx not installed")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(1)


BASE = "/root/dabt-curated"

OUTPUTS = {
    "chapter": "/tmp/chapter_fingerprints.csv",
    "kristen_topic": "/tmp/kristen_topic_fingerprints.csv",
    "kristen_mini": "/tmp/kristen_mini_fingerprints.csv",
    "past_exam": "/tmp/past_exam_fingerprints.csv",
}


def clean_text(text):
    """Lowercase, strip punctuation (keep only letters, digits, spaces), collapse whitespace."""
    if not text:
        return ""
    text = text.lower()
    # strip punctuation: keep a-z, 0-9, whitespace
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    # collapse whitespace
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def make_fingerprint(text, length=80):
    """Create fingerprint: first `length` chars after cleaning."""
    cleaned = clean_text(text)
    return cleaned[:length]


def extract_from_docx(filepath):
    """Extract paragraphs from a .docx file. Returns list of (paragraph_num, text)."""
    try:
        doc = Document(filepath)
    except Exception as e:
        print(f"  ERROR opening docx {filepath}: {e}")
        return []
    
    results = []
    for i, para in enumerate(doc.paragraphs):
        text = para.text.strip()
        if text:
            results.append((i + 1, text))
    return results


def extract_from_xlsx(filepath):
    """Extract cell text from .xlsx file. Returns list of (row_num_str, text)."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR opening xlsx {filepath}: {e}")
        return []
    
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, cell in enumerate(row):
                if cell is not None:
                    text = str(cell).strip()
                    if text:
                        label = f"{sheet_name} R{row_idx}C{col_idx+1}"
                        results.append((label, text))
    wb.close()
    return results


def should_skip_kristen(filename):
    """Skip files with 'answer' or 'explanation' (or misspelling 'explaination') in filename (case-insensitive)."""
    lower = filename.lower()
    if 'answer' in lower:
        return True
    # Match both 'explanation' and misspelled 'explaination' / 'explainations'
    if 'explana' in lower or 'explaina' in lower:
        return True
    return False


def process_items(items, csv_path, label):
    """
    items: list of (filepath, source_label)
    Extract questions from each file, compute fingerprints, write CSV.
    """
    rows = []
    for filepath, src_label in items:
        print(f"  Processing: {src_label}")
        if filepath.endswith('.docx'):
            entries = extract_from_docx(filepath)
        elif filepath.endswith('.xlsx'):
            entries = extract_from_xlsx(filepath)
        else:
            print(f"    Skipping unknown format: {filepath}")
            continue
        
        for num, text in entries:
            fp = make_fingerprint(text)
            rows.append({
                'File': src_label,
                'QuestionNumber': num,
                'Fingerprint': fp,
                'FullText': text[:200],  # truncate long text for readability
            })
        
        print(f"    Found {len(entries)} text segments -> {len(rows)} rows so far")
    
    # Write CSV
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['File', 'QuestionNumber', 'Fingerprint', 'FullText'])
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Wrote {len(rows)} rows to {csv_path}")
    return rows


def main():
    # =========================================================
    # A. Chapter Tests
    # =========================================================
    print("\n" + "=" * 60)
    print("A. Chapter Tests (all .docx files in Tests/)")
    test_dir = os.path.join(BASE, "Chapter_Tests", "Tests")
    chapter_files = sorted(glob.glob(os.path.join(test_dir, "*.docx")))
    print(f"  Found {len(chapter_files)} .docx files")
    chapter_items = [(f, os.path.basename(f)) for f in chapter_files]
    process_items(chapter_items, OUTPUTS["chapter"], "Chapter Tests")

    # =========================================================
    # B. Kristen Topic Tests
    # =========================================================
    print("\n" + "=" * 60)
    print("B. Kristen Topic Tests (skip files with 'answer' or 'explanation')")
    topic_dir = os.path.join(BASE, "Practice_Tests_by_Topic", "Kristen_Topic_Tests")
    all_topic_files = sorted(glob.glob(os.path.join(topic_dir, "*.docx")))
    topic_files = [f for f in all_topic_files if not should_skip_kristen(os.path.basename(f))]
    print(f"  Found {len(all_topic_files)} total, {len(topic_files)} after filtering")
    for f in sorted(set(all_topic_files) - set(topic_files)):
        print(f"    Skipped: {os.path.basename(f)}")
    topic_items = [(f, os.path.basename(f)) for f in topic_files]
    process_items(topic_items, OUTPUTS["kristen_topic"], "Kristen Topic Tests")

    # =========================================================
    # C. Kristen Mini Exams
    # =========================================================
    print("\n" + "=" * 60)
    print("C. Kristen Mini Exams (skip files with 'answer' or 'explanation')")
    mini_dir = os.path.join(BASE, "Practice_Exams", "Kristen_Mini_Exams")
    all_mini_files = sorted(glob.glob(os.path.join(mini_dir, "*.docx")))
    mini_files = [f for f in all_mini_files if not should_skip_kristen(os.path.basename(f))]
    print(f"  Found {len(all_mini_files)} total, {len(mini_files)} after filtering")
    for f in sorted(set(all_mini_files) - set(mini_files)):
        print(f"    Skipped: {os.path.basename(f)}")
    mini_items = [(f, os.path.basename(f)) for f in mini_files]
    process_items(mini_items, OUTPUTS["kristen_mini"], "Kristen Mini Exams")

    # =========================================================
    # D. Past ABT Exams
    # =========================================================
    print("\n" + "=" * 60)
    print("D. Past ABT Exams (.docx files and compiled recert exams .xlsx)")
    past_dir = os.path.join(BASE, "Practice_Exams", "Past_ABT_Exams")
    
    past_items = []
    
    # Find .docx files (including in subdirectories)
    docx_files = sorted(glob.glob(os.path.join(past_dir, "**", "*.docx"), recursive=True))
    # Skip any that look like answer-only files (filename contains "answer" but not as part of a question)
    for f in docx_files:
        basename = os.path.basename(f)
        rel = os.path.relpath(f, past_dir)
        # Skip if in Recert_Discussion_Slides and filename suggests answer-only
        if 'answer' in basename.lower() and 'question' not in basename.lower():
            print(f"    Skipped (answer-only?): {rel}")
            continue
        past_items.append((f, rel))
    
    # Find the compiled recert exams .xlsx
    xlsx_files = sorted(glob.glob(os.path.join(past_dir, "*.xlsx")))
    for f in xlsx_files:
        rel = os.path.relpath(f, past_dir)
        past_items.append((f, rel))
    
    print(f"  Found {len(past_items)} files to process")
    for f, rel in past_items:
        print(f"    {rel}")
    
    process_items(past_items, OUTPUTS["past_exam"], "Past ABT Exams")

    print("\n" + "=" * 60)
    print("ALL DONE.")
    for key, path in OUTPUTS.items():
        if os.path.exists(path):
            size = os.path.getsize(path)
            print(f"  {path}: {size} bytes")


if __name__ == "__main__":
    main()
