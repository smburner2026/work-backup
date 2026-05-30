#!/usr/bin/env python3
"""
Build comprehensive DABT SQLite database from all source files — v2.
Fixes: matching question handling, PDF dedup, Past ABT answer parsing, domain fallback.
"""
import sqlite3
import csv
import openpyxl
import os
import sys
import re
import hashlib
from datetime import datetime

# === Configuration ===
DB_PATH = "/root/work/dabt/dabt-tutor/reference/data/dabt.db"
LOG_PATH = "/root/dabt_import_log.txt"
STATS_PATH = "/root/dabt_db_stats.txt"
EXISTING_XLSX = "/root/work/dabt/dabt-tutor/reference/data/DABT_Practice_Questions_Database.xlsx"
CLASSIFICATION_CSV = "/root/work/dabt/dabt-tutor/reference/data/question_classifications.csv"
CSV_2000Q = "/tmp/dabt_extract_2000q.csv"
CSV_CHAPTER = "/tmp/dabt_extract_chapter.csv"
CSV_MINI = "/tmp/dabt_extract_mini.csv"
CSV_TOPIC = "/tmp/dabt_extract_topic.csv"
PAST_ABT_XLSX = "/root/dabt-curated/Practice_Exams/Past_ABT_Exams/2008-2014_Compiled_Recert_Exams.xlsx"
PAST_ABT_DIR = "/root/dabt-curated/Practice_Exams/Past_ABT_Exams"

# Domain by topic mapping (primary exact matches)
DOMAIN_BY_TOPIC = {
    "General Principles & Concepts": "Domain I",
    "General Toxicology": "Domain I",
    "Toxicokinetics / ADME": "Domain I",
    "Toxicokinetics": "Domain I",
    "ADME": "Domain I",
    "Biotransformation / Metabolism": "Domain II",
    "Mechanisms of Toxicity": "Domain II",
    "Carcinogenesis & Mutagenesis": "Domain II",
    "Carcinogenesis": "Domain II",
    "Genotoxicity / DNA Damage": "Domain II",
    "Genotoxicity": "Domain II",
    "Risk Assessment & Regulatory": "Domain III",
    "Risk Assessment": "Domain III",
    "Regulatory Toxicology": "Domain III",
    "Liver / Hepatotoxicity": "Domain IV",
    "Kidney / Nephrotoxicity": "Domain IV",
    "Hematology & Blood Toxicity": "Domain IV",
    "Immunotoxicology / Allergy": "Domain IV",
    "Lung / Pulmonary Toxicity": "Domain IV",
    "Nervous System / Neurotoxicity": "Domain IV",
    "Eye / Ocular Toxicity": "Domain IV",
    "Cardiovascular Toxicity": "Domain IV",
    "Skin / Dermatotoxicity": "Domain IV",
    "Endocrine Toxicology": "Domain IV",
    "Reproductive & Developmental Toxicity": "Domain IV",
    "Metals & Metalloids": "Domain IV",
    "Solvents & Hydrocarbons": "Domain IV",
    "Pesticides – Insecticides": "Domain IV",
    "Air Pollution & Particulates": "Domain IV",
    "Plant Toxins": "Domain IV",
    "Food Additives, Cosmetics & GRAS": "Domain IV",
    "Drugs & Therapeutics – Toxicology": "Domain IV",
    "Alcohols & Methanol/Ethanol": "Domain IV",
    "Radiation / UV / Ionizing": "Domain IV",
    "Animal & Microbial Venoms / Toxins": "Domain IV",
    "Mycotoxins": "Domain IV",
    "Gases – Asphyxiants & Irritants": "Domain IV",
}

# Additional keyword-based domain inference for untagged questions
KEYWORD_DOMAIN = {
    # Domain I - General
    "general principle": "Domain I", "general toxicology": "Domain I",
    "toxicokinetic": "Domain I", "adme": "Domain I", "absorption": "Domain I",
    "distribution": "Domain I", "excretion": "Domain I",
    "dose-response": "Domain I", "bioassay": "Domain I", "statistics": "Domain I",
    "study design": "Domain I", "ld50": "Domain I", "lc50": "Domain I",
    "noael": "Domain I", "noel": "Domain I", "loael": "Domain I",
    "benchmark dose": "Domain I", "bmd": "Domain I",
    
    # Domain II - Mechanisms
    "biotransformation": "Domain II", "metabolism": "Domain II", "phase i": "Domain II",
    "phase ii": "Domain II", "cyp": "Domain II", "p450": "Domain II",
    "mechanism": "Domain II", "mode of action": "Domain II", "moa": "Domain II",
    "apoptosis": "Domain II", "necrosis": "Domain II", "oxidative stress": "Domain II",
    "carcinogen": "Domain II", "mutagen": "Domain II", "genotox": "Domain II",
    "dna damage": "Domain II", "dna repair": "Domain II", "ames": "Domain II",
    "micronucleus": "Domain II", "chromosomal": "Domain II",
    
    # Domain III - Risk Assessment
    "risk assess": "Domain III", "hazard": "Domain III", "exposure": "Domain III",
    "risk characterization": "Domain III", "risk management": "Domain III",
    "uncertainty factor": "Domain III", "margin of safety": "Domain III",
    "mos": "Domain III", "reference dose": "Domain III", "rfd": "Domain III",
    "acceptable daily": "Domain III", "adi": "Domain III", "tolerable": "Domain III",
    "regulatory": "Domain III", "epa": "Domain III", "fda": "Domain III",
    "niosh": "Domain III", "osha": "Domain III", "classification": "Domain III",
    "labeling": "Domain III", "ghs": "Domain III",
    
    # Domain IV - Applied (specific topics)
    "liver": "Domain IV", "hepatotox": "Domain IV", "hepatic": "Domain IV",
    "kidney": "Domain IV", "renal": "Domain IV", "nephrotox": "Domain IV",
    "blood": "Domain IV", "hematol": "Domain IV", "erythrocyte": "Domain IV",
    "anemia": "Domain IV", "methemoglobin": "Domain IV",
    "immuno": "Domain IV", "allerg": "Domain IV", "hypersensitivity": "Domain IV",
    "lung": "Domain IV", "pulmonary": "Domain IV", "respirat": "Domain IV",
    "inhalation": "Domain IV", "asthma": "Domain IV",
    "nervous": "Domain IV", "neurotox": "Domain IV", "brain": "Domain IV",
    "neuron": "Domain IV", "synapse": "Domain IV",
    "eye": "Domain IV", "ocular": "Domain IV", "vision": "Domain IV",
    "cardio": "Domain IV", "heart": "Domain IV", "arrhythmia": "Domain IV",
    "skin": "Domain IV", "dermat": "Domain IV", "dermal": "Domain IV",
    "endocrine": "Domain IV", "hormone": "Domain IV", "thyroid": "Domain IV",
    "repro": "Domain IV", "develop": "Domain IV", "teratogen": "Domain IV",
    "fetal": "Domain IV", "fertility": "Domain IV", "pregnan": "Domain IV",
    "metal": "Domain IV", "mercury": "Domain IV", "lead": "Domain IV",
    "cadmium": "Domain IV", "arsenic": "Domain IV", "chromium": "Domain IV",
    "solvent": "Domain IV", "hydrocarbon": "Domain IV", "benzene": "Domain IV",
    "pesticide": "Domain IV", "insecticide": "Domain IV", "organophosphate": "Domain IV",
    "air pollution": "Domain IV", "particulate": "Domain IV", "ozone": "Domain IV",
    "plant toxin": "Domain IV", "mushroom": "Domain IV", "phyto": "Domain IV",
    "food additive": "Domain IV", "cosmetic": "Domain IV", "gras": "Domain IV",
    "drug": "Domain IV", "therapeutic": "Domain IV", "acetaminophen": "Domain IV",
    "alcohol": "Domain IV", "ethanol": "Domain IV", "methanol": "Domain IV",
    "radiation": "Domain IV", "uv": "Domain IV", "ionizing": "Domain IV",
    "venom": "Domain IV", "snake": "Domain IV", "spider": "Domain IV",
    "mycotoxin": "Domain IV", "aflatoxin": "Domain IV",
    "gas": "Domain IV", "asphyxiant": "Domain IV", "carbon monoxide": "Domain IV",
    "cyanide": "Domain IV", "animal toxin": "Domain IV", "microbial": "Domain IV",
}

log_lines = []

def log(msg):
    print(msg)
    log_lines.append(msg)

def text_or_none(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def text_or_empty(val):
    if val is None:
        return ""
    return str(val).strip()

def create_tables(conn):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS source_files (
        id INTEGER PRIMARY KEY,
        bank_name TEXT NOT NULL,
        filename TEXT NOT NULL,
        format_type TEXT,
        year TEXT,
        description TEXT
    );

    CREATE TABLE IF NOT EXISTS questions (
        id TEXT PRIMARY KEY,
        question_text TEXT NOT NULL,
        correct_answer_letter TEXT,
        correct_answer_text TEXT,
        explanation TEXT,
        source_file_id INTEGER,
        question_number_in_source INTEGER,
        bloom_level TEXT,
        FOREIGN KEY (source_file_id) REFERENCES source_files(id)
    );

    CREATE TABLE IF NOT EXISTS answer_options (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        question_id TEXT NOT NULL,
        option_letter TEXT NOT NULL,
        option_text TEXT NOT NULL,
        FOREIGN KEY (question_id) REFERENCES questions(id)
    );

    CREATE TABLE IF NOT EXISTS question_topics (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        question_id TEXT NOT NULL,
        topic TEXT NOT NULL,
        FOREIGN KEY (question_id) REFERENCES questions(id)
    );

    CREATE TABLE IF NOT EXISTS question_domains (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        question_id TEXT NOT NULL,
        domain TEXT NOT NULL,
        sub_domain TEXT,
        task TEXT,
        confidence TEXT,
        FOREIGN KEY (question_id) REFERENCES questions(id)
    );

    CREATE TABLE IF NOT EXISTS match_pairs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        question_id TEXT NOT NULL,
        term TEXT NOT NULL,
        match_answer TEXT NOT NULL,
        FOREIGN KEY (question_id) REFERENCES questions(id)
    );

    CREATE INDEX IF NOT EXISTS idx_q_domain ON question_domains(domain);
    CREATE INDEX IF NOT EXISTS idx_q_topic ON question_topics(topic);
    CREATE INDEX IF NOT EXISTS idx_q_source ON questions(source_file_id);
    """)
    conn.commit()

def parse_options(row, question_id, cursor):
    """Parse A through H columns into answer_options table."""
    option_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    count = 0
    for letter in option_letters:
        val = text_or_none(row.get(letter))
        if val:
            cursor.execute(
                "INSERT INTO answer_options (question_id, option_letter, option_text) VALUES (?, ?, ?)",
                (question_id, letter, val)
            )
            count += 1
    return count

def parse_topics(row, question_id, cursor):
    """Parse Topic (Primary) and All Topics into question_topics table."""
    primary = text_or_none(row.get('Topic (Primary)'))
    all_topics = text_or_none(row.get('All Topics'))
    
    seen = set()
    if primary and primary not in seen:
        cursor.execute("INSERT INTO question_topics (question_id, topic) VALUES (?, ?)",
                      (question_id, primary))
        seen.add(primary)
    
    if all_topics:
        for t in re.split(r'[;,|]', all_topics):
            t = t.strip()
            if t and t not in seen:
                cursor.execute("INSERT INTO question_topics (question_id, topic) VALUES (?, ?)",
                              (question_id, t))
                seen.add(t)

def load_classifications():
    classifications = {}
    if not os.path.exists(CLASSIFICATION_CSV):
        log(f"WARNING: Classification CSV not found at {CLASSIFICATION_CSV}")
        return classifications
    
    with open(CLASSIFICATION_CSV, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            qid = row.get('ID', '').strip()
            if qid:
                classifications[qid] = {
                    'domain': text_or_none(row.get('Domain')),
                    'sub_domain': text_or_none(row.get('Sub-Domain')),
                    'task': text_or_none(row.get('Task')),
                    'bloom_level': text_or_none(row.get('Bloom Level')),
                    'domain_confidence': text_or_none(row.get('Domain Confidence')),
                    'bloom_confidence': text_or_none(row.get('Bloom Confidence')),
                }
    log(f"Loaded {len(classifications)} classifications from CSV")
    return classifications

def assign_domains(conn, classifications):
    cursor = conn.cursor()
    
    cursor.execute("SELECT id FROM questions")
    all_questions = [row[0] for row in cursor.fetchall()]
    
    assigned_count = 0
    topic_mapped_count = 0
    keyword_mapped_count = 0
    
    for qid in all_questions:
        # 1. Check classification
        if qid in classifications:
            cls = classifications[qid]
            if cls['domain']:
                cursor.execute(
                    "INSERT INTO question_domains (question_id, domain, sub_domain, task, confidence) VALUES (?, ?, ?, ?, ?)",
                    (qid, cls['domain'], cls['sub_domain'], cls['task'], cls['domain_confidence'])
                )
                assigned_count += 1
                continue
        
        # 2. Try topic → domain mapping
        cursor.execute("SELECT topic FROM question_topics WHERE question_id = ?", (qid,))
        topics = [row[0] for row in cursor.fetchall()]
        
        domain_found = False
        for topic in topics:
            domain = DOMAIN_BY_TOPIC.get(topic)
            if not domain:
                for k, v in DOMAIN_BY_TOPIC.items():
                    if k.lower() == topic.lower():
                        domain = v
                        break
            if not domain:
                for k, v in DOMAIN_BY_TOPIC.items():
                    if k.split('/')[0].strip().lower() == topic.lower():
                        domain = v
                        break
            if not domain:
                # Try partial match
                t_lower = topic.lower()
                for k, v in DOMAIN_BY_TOPIC.items():
                    if k.lower() in t_lower or t_lower in k.lower():
                        domain = v
                        break
            
            if domain:
                cursor.execute(
                    "INSERT OR IGNORE INTO question_domains (question_id, domain) VALUES (?, ?)",
                    (qid, domain)
                )
                topic_mapped_count += 1
                domain_found = True
                break
        
        if domain_found:
            continue
        
        # 3. Try keyword inference from question text
        cursor.execute("SELECT question_text, correct_answer_text, explanation FROM questions WHERE id = ?", (qid,))
        qrow = cursor.fetchone()
        if qrow:
            combined = ' '.join(filter(None, qrow)).lower()
            for keyword, domain in sorted(KEYWORD_DOMAIN.items(), key=lambda x: -len(x[0])):
                if keyword in combined:
                    cursor.execute(
                        "INSERT OR IGNORE INTO question_domains (question_id, domain, confidence) VALUES (?, ?, ?)",
                        (qid, domain, 'low')
                    )
                    keyword_mapped_count += 1
                    domain_found = True
                    break
    
    conn.commit()
    log(f"Domain assignment: {assigned_count} from classification, {topic_mapped_count} from topic mapping, {keyword_mapped_count} from keyword")

def get_question_text_hash(question_text):
    normalized = ' '.join(question_text.lower().split())
    return hashlib.md5(normalized.encode('utf-8')).hexdigest()

def is_matching_type_question(row):
    """Detect if a row is a matching-type question (term only, shared options)."""
    question = text_or_none(row.get('Question'))
    if not question:
        return False
    
    # Short question text (single term or phrase) with no answer options
    has_opts = any(text_or_none(row.get(l)) for l in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'])
    if not has_opts and len(question.split()) <= 5:
        return True
    
    # Or has exactly 1 option column filled (index-style matching)
    opt_count = sum(1 for l in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'] if text_or_none(row.get(l)))
    if opt_count == 1 and len(question.split()) <= 6:
        return True
    
    return False

def load_existing_xlsx(conn, source_file_id):
    """Load the existing 446 questions from the xlsx file."""
    cursor = conn.cursor()
    
    wb = openpyxl.load_workbook(EXISTING_XLSX)
    ws = wb['Questions']
    
    headers = [ws.cell(1, c).value for c in range(1, 19)]
    
    inserted = 0
    existing_hashes = {}
    
    for r in range(2, ws.max_row + 1):
        row_data = {}
        for c in range(1, 19):
            row_data[headers[c-1]] = ws.cell(r, c).value
        
        qid = text_or_none(row_data.get('ID'))
        if not qid:
            continue
        
        question_text = text_or_none(row_data.get('Question'))
        if not question_text:
            continue
        
        correct_letter = text_or_none(row_data.get('Correct Answer'))
        correct_text = text_or_none(row_data.get('Correct Answer Text'))
        explanation = text_or_none(row_data.get('Explanation'))
        qnum = text_or_none(row_data.get('Question #'))
        
        cursor.execute(
            "INSERT INTO questions (id, question_text, correct_answer_letter, correct_answer_text, explanation, source_file_id, question_number_in_source) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (qid, question_text, correct_letter, correct_text, explanation, source_file_id, qnum)
        )
        
        parse_options(row_data, qid, cursor)
        parse_topics(row_data, qid, cursor)
        
        qhash = get_question_text_hash(question_text)
        existing_hashes[qhash] = qid
        
        inserted += 1
        if inserted % 100 == 0:
            conn.commit()
    
    conn.commit()
    log(f"Loaded {inserted} questions from existing xlsx")
    return existing_hashes

def load_csv_data(filepath, conn, source_file_id, start_id_num, existing_hashes):
    """
    Load questions from a CSV file.
    Handles matching-type questions separately.
    """
    cursor = conn.cursor()
    inserted = 0
    skipped = 0
    match_inserted = 0
    
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            question_text = text_or_none(row.get('Question'))
            if not question_text:
                skipped += 1
                continue
            
            # Dedup check
            qhash = get_question_text_hash(question_text)
            if qhash in existing_hashes:
                skipped += 1
                continue
            
            # Check if matching-type question
            if is_matching_type_question(row):
                # Store as match_pairs: the question text is the term,
                # and correct_answer tells what it matches to
                correct_letter = text_or_none(row.get('Correct Answer'))
                if correct_letter:
                    qid = f"DABT-{start_id_num:04d}"
                    start_id_num += 1
                    
                    # Find the option text if present in any column
                    match_text = None
                    for l in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                        val = text_or_none(row.get(l))
                        if val:
                            match_text = val
                            break
                    
                    cursor.execute(
                        "INSERT INTO questions (id, question_text, correct_answer_letter, correct_answer_text, source_file_id, question_number_in_source) VALUES (?, ?, ?, ?, ?, ?)",
                        (qid, question_text, correct_letter, match_text, source_file_id, text_or_none(row.get('Question #')))
                    )
                    
                    # Store in match_pairs
                    cursor.execute(
                        "INSERT INTO match_pairs (question_id, term, match_answer) VALUES (?, ?, ?)",
                        (qid, question_text, correct_letter)
                    )
                    
                    # Add the option text as an answer option if available
                    if match_text:
                        cursor.execute(
                            "INSERT INTO answer_options (question_id, option_letter, option_text) VALUES (?, ?, ?)",
                            (qid, correct_letter, match_text)
                        )
                    
                    parse_topics(row, qid, cursor)
                    existing_hashes[qhash] = qid
                    match_inserted += 1
                else:
                    # No correct answer either — skip incomplete row
                    skipped += 1
                continue
            
            # Normal question
            qid = f"DABT-{start_id_num:04d}"
            start_id_num += 1
            
            correct_letter = text_or_none(row.get('Correct Answer'))
            correct_text = text_or_none(row.get('Correct Answer Text'))
            explanation = text_or_none(row.get('Explanation'))
            qnum = text_or_none(row.get('Question #'))
            
            cursor.execute(
                "INSERT INTO questions (id, question_text, correct_answer_letter, correct_answer_text, explanation, source_file_id, question_number_in_source) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (qid, question_text, correct_letter, correct_text, explanation, source_file_id, qnum)
            )
            
            parse_options(row, qid, cursor)
            parse_topics(row, qid, cursor)
            
            existing_hashes[qhash] = qid
            inserted += 1
            
            if (inserted + match_inserted) % 100 == 0:
                conn.commit()
    
    conn.commit()
    log(f"{os.path.basename(filepath)}: inserted {inserted} normal + {match_inserted} matching, skipped {skipped}")
    return start_id_num

def parse_past_abt_exams(conn, source_file_id, start_id_num, existing_hashes):
    """Parse the Past ABT Exams xlsx file with embedded question/answer format."""
    cursor = conn.cursor()
    inserted = 0
    skipped = 0
    
    wb = openpyxl.load_workbook(PAST_ABT_XLSX)
    ws = wb['Sheet1']
    
    for r in range(1, ws.max_row + 1):
        col_a = text_or_empty(ws.cell(r, 1).value)
        col_b = text_or_empty(ws.cell(r, 2).value)
        
        if not col_a:
            continue
        
        lines = col_a.split('\n')
        
        # Find where answer options start
        option_start = None
        for i, line in enumerate(lines):
            line_stripped = line.strip()
            if re.match(r'^[A-H][\).:]\s', line_stripped):
                option_start = i
                break
        
        if option_start is None:
            question_text = col_a
            options = {}
        else:
            question_text = '\n'.join(lines[:option_start]).strip()
            options = {}
            for line in lines[option_start:]:
                line = line.strip()
                m = re.match(r'^([A-H])[\).:]\s*(.*)', line)
                if m:
                    options[m.group(1)] = m.group(2).strip()
        
        # Parse answer and explanation from Col B
        correct_letter = None
        explanation = None
        
        if col_b:
            col_b = col_b.strip()
            
            # Try multiple answer patterns
            patterns = [
                r'(?:Correct\s*Answer|Answer)\s*[=:]\s*([A-H])',
                r'Answer\s+is\s+([A-H])[\.\s]',
                r'^([A-H])[\.\)]\s*(?:is|correct)',
                r'^The\s+correct\s+answer\s+is\s+([A-H])',
            ]
            for pat in patterns:
                m = re.search(pat, col_b, re.IGNORECASE)
                if m:
                    correct_letter = m.group(1)
                    break
            
            # Also check format: "3. Correct Answer = A"
            if not correct_letter:
                m = re.search(r'(?:^|\n)\d+[\.\)]\s*Correct\s*Answer\s*[=:]\s*([A-H])', col_b, re.IGNORECASE)
                if m:
                    correct_letter = m.group(1)
            
            # Extract explanation (everything after the answer statement)
            explanation = re.sub(
                r'^(?:\d+[\.\)]\s*)?(?:Correct\s*Answer|Answer)\s*[=:]\s*[A-H]\s*\n?\s*',
                '', col_b, flags=re.IGNORECASE
            )
            explanation = re.sub(r'^Answer\s+is\s+[A-H][\.\s]*', '', explanation, flags=re.IGNORECASE)
            explanation = re.sub(r'^[A-H][\.\)]\s+', '', explanation)
            explanation = re.sub(r'^\d+[\.\)]\s*', '', explanation)
            explanation = explanation.strip()
        
        if not question_text:
            skipped += 1
            continue
        
        # Dedup
        qhash = get_question_text_hash(question_text)
        if qhash in existing_hashes:
            skipped += 1
            continue
        
        qid = f"DABT-{start_id_num:04d}"
        start_id_num += 1
        
        cursor.execute(
            "INSERT INTO questions (id, question_text, correct_answer_letter, correct_answer_text, explanation, source_file_id, question_number_in_source) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (qid, question_text, correct_letter, None, explanation, source_file_id, r)
        )
        
        for letter in sorted(options.keys()):
            cursor.execute(
                "INSERT INTO answer_options (question_id, option_letter, option_text) VALUES (?, ?, ?)",
                (qid, letter, options[letter])
            )
        
        # Try to infer topics from question text
        infer_topic_from_text(question_text, qid, cursor)
        
        existing_hashes[qhash] = qid
        inserted += 1
        
        if inserted % 100 == 0:
            conn.commit()
    
    conn.commit()
    log(f"Past ABT Exams (2008-2014): loaded {inserted}, skipped {skipped} (dedup or empty)")
    return start_id_num

def infer_topic_from_text(question_text, qid, cursor):
    """Try to infer topic from question text using keyword matching."""
    keyword_topics = {
        "mercury|methylmercury|lead|cadmium|arsenic|chromium|nickel|chelating|chelation": "Metals & Metalloids",
        "carcinogen|mutagen|genotox|dna damage|ames test|micronucleus|sister chromatid": "Carcinogenesis & Mutagenesis",
        "liver|hepatotox|hepatic|biliary|cholestasis|hepatocyte": "Liver / Hepatotoxicity",
        "kidney|renal|nephrotox|glomerular|tubular|nephron": "Kidney / Nephrotoxicity",
        "neurotox|nervous|brain|neuron|synapse|neuropathy|axon|myelin": "Nervous System / Neurotoxicity",
        "skin|dermatotox|dermal|epidermis|contact dermatitis": "Skin / Dermatotoxicity",
        "lung|pulmonary|respirat|inhalation|asthma|bronch|alveoli": "Lung / Pulmonary Toxicity",
        "heart|cardio|cardiovascular|cardiotox|arrhythmia|myocardial": "Cardiovascular Toxicity",
        "reproduc|develop|teratogen|fetotox|embryotox|fertility|placenta": "Reproductive & Developmental Toxicity",
        "endocrine|hormone|thyroid|estrogen|androgen|steroid|adrenal": "Endocrine Toxicology",
        "immunotox|allerg|hypersensitivity|autoimmune|lymphocyte|antibody": "Immunotoxicology / Allergy",
        "blood|hematol|erythrocyte|hemoglobin|anemia|methemoglobin|coagulation": "Hematology & Blood Toxicity",
        "eye|ocular|retina|cataract|corneal|vision|ophthalmic": "Eye / Ocular Toxicity",
        "pesticide|insecticide|herbicide|organophosphate|carbamate|pyrethroid": "Pesticides – Insecticides",
        "solvent|hydrocarbon|benzene|toluene|xylene|acetone|chloroform|methylene": "Solvents & Hydrocarbons",
        "risk assess|hazard|exposure|dose-response|noael|loael|margin of safety|uncertainty factor": "Risk Assessment & Regulatory",
        "toxicokinetic|adme|absorption|distribution|metabolism|excretion|bioavail|half-life": "Toxicokinetics / ADME",
        "biotransform|metabolism|phase i|phase ii|cyp|cytochrome|p450|conjugation": "Biotransformation / Metabolism",
        "mechanism|mode of action|aop|toxicant|receptor|signaling|apoptosis|necrosis": "Mechanisms of Toxicity",
        "alcohol|ethanol|methanol|ethylene glycol|isopropanol": "Alcohols & Methanol/Ethanol",
        "radiation|uv|ionizing|radionuclide|radioactive": "Radiation / UV / Ionizing",
        "plant toxin|phyto|herbal|mushroom|amanita|ricin": "Plant Toxins",
        "venom|snake|spider|scorpion|toxin|animal|microbial|botulinum|tetrodotoxin": "Animal & Microbial Venoms / Toxins",
        "mycotoxin|aflatoxin|ochratoxin|fumonisin|ergot": "Mycotoxins",
        "gas|asphyxiant|irritant|carbon monoxide|cyanide|hydrogen sulfide|chlorine|phosgene": "Gases – Asphyxiants & Irritants",
        "food additive|cosmetic|gras|preservative|additive|contaminant": "Food Additives, Cosmetics & GRAS",
        "drug|therapeutic|acetaminophen|paracetamol|aspirin|digitalis|chemotherapy": "Drugs & Therapeutics – Toxicology",
        "air pollution|particulate|pm2.5|pm10|ozone|smog|diesel": "Air Pollution & Particulates",
        "principle|concept|general|toxicology": "General Principles & Concepts",
    }
    
    qtext_lower = question_text.lower()
    for pattern, topic in keyword_topics.items():
        if re.search(pattern, qtext_lower):
            cursor.execute(
                "INSERT OR IGNORE INTO question_topics (question_id, topic) VALUES (?, ?)",
                (qid, topic)
            )
            return

def parse_pdf_files(conn, source_file_id, start_id_num, existing_hashes):
    """Parse PDF files — skip files that are duplicates by size."""
    cursor = conn.cursor()
    inserted = 0
    
    try:
        import fitz
    except ImportError:
        log("pymupdf not available, skipping PDF parsing")
        return start_id_num
    
    pdf_files = sorted([f for f in os.listdir(PAST_ABT_DIR) if f.endswith('.pdf')])
    
    # Group by file size to detect duplicates
    size_to_files = {}
    for f in pdf_files:
        fp = os.path.join(PAST_ABT_DIR, f)
        sz = os.path.getsize(fp)
        size_to_files.setdefault(sz, []).append(f)
    
    # Report duplicates
    for sz, files in size_to_files.items():
        if len(files) > 1:
            log(f"  Duplicate sizes ({sz} bytes): {files}")
    
    # Dedup: skip files whose size was already processed
    processed_sizes = set()
    
    for pdf_file in pdf_files:
        pdf_path = os.path.join(PAST_ABT_DIR, pdf_file)
        sz = os.path.getsize(pdf_path)
        
        if sz in processed_sizes:
            log(f"  SKIPPING (duplicate): {pdf_file} (same size as already processed)")
            continue
        
        log(f"Parsing PDF: {pdf_file}")
        
        try:
            doc = fitz.open(pdf_path)
            full_text = ""
            for page in doc:
                full_text += page.get_text() + "\n\n"
            doc.close()
        except Exception as e:
            log(f"  Error reading PDF {pdf_file}: {e}")
            continue
        
        lines = full_text.split('\n')
        
        questions_found = []
        current_question = []
        in_question = False
        
        for line in lines:
            stripped = line.strip()
            if not stripped:
                if in_question:
                    questions_found.append('\n'.join(current_question))
                    current_question = []
                    in_question = False
                continue
            
            if re.match(r'^\d+[\.\)]\s', stripped) and len(stripped) > 5:
                if in_question:
                    questions_found.append('\n'.join(current_question))
                current_question = [stripped]
                in_question = True
            elif in_question:
                current_question.append(stripped)
        
        if in_question and current_question:
            questions_found.append('\n'.join(current_question))
        
        pdf_inserted = 0
        for q_block in questions_found:
            lines = q_block.split('\n')
            
            option_start = None
            for i, line in enumerate(lines):
                if re.match(r'^[A-H][\).:]\s', line.strip()):
                    option_start = i
                    break
            
            if option_start is None:
                continue
            
            question_text = '\n'.join(lines[:option_start]).strip()
            question_text = re.sub(r'^\d+[\.\)]\s*', '', question_text)
            
            if not question_text or len(question_text) < 10:
                continue
            
            options = {}
            for line in lines[option_start:]:
                line = line.strip()
                m = re.match(r'^([A-H])[\).:]\s*(.*)', line)
                if m:
                    options[m.group(1)] = m.group(2).strip()
            
            if len(options) < 2:
                continue
            
            # Dedup
            qhash = get_question_text_hash(question_text)
            if qhash in existing_hashes:
                continue
            
            qid = f"DABT-{start_id_num:04d}"
            start_id_num += 1
            
            cursor.execute(
                "INSERT INTO questions (id, question_text, source_file_id) VALUES (?, ?, ?)",
                (qid, question_text, source_file_id)
            )
            
            for letter in sorted(options.keys()):
                cursor.execute(
                    "INSERT INTO answer_options (question_id, option_letter, option_text) VALUES (?, ?, ?)",
                    (qid, letter, options[letter])
                )
            
            infer_topic_from_text(question_text, qid, cursor)
            
            existing_hashes[qhash] = qid
            pdf_inserted += 1
            inserted += 1
            
            if inserted % 100 == 0:
                conn.commit()
        
        conn.commit()
        processed_sizes.add(sz)
        log(f"  Inserted {pdf_inserted} questions from {pdf_file}")
    
    return start_id_num

def parse_docx_files(conn, source_file_id, start_id_num, existing_hashes):
    """Parse .docx files."""
    cursor = conn.cursor()
    inserted = 0
    
    try:
        from docx import Document
    except ImportError:
        log("python-docx not available, skipping docx parsing")
        return start_id_num
    
    for root_dir, dirs, files in os.walk(PAST_ABT_DIR):
        for filename in files:
            if not filename.endswith('.docx'):
                continue
            filepath = os.path.join(root_dir, filename)
            # Skip archive/template files
            if '~$' in filename:
                continue
            
            log(f"Parsing DOCX: {filepath}")
            
            try:
                doc = Document(filepath)
                full_text = [para.text for para in doc.paragraphs]
                text = '\n'.join(full_text)
            except Exception as e:
                log(f"  Error: {e}")
                continue
            
            lines = text.split('\n')
            questions_found = []
            current_question = []
            in_question = False
            
            for line in lines:
                stripped = line.strip()
                if not stripped:
                    if in_question:
                        questions_found.append('\n'.join(current_question))
                        current_question = []
                        in_question = False
                    continue
                
                if re.match(r'^\d+[\.\)]\s', stripped) and len(stripped) > 5:
                    if in_question:
                        questions_found.append('\n'.join(current_question))
                    current_question = [stripped]
                    in_question = True
                elif in_question:
                    current_question.append(stripped)
            
            if in_question and current_question:
                questions_found.append('\n'.join(current_question))
            
            docx_inserted = 0
            for q_block in questions_found:
                lines = q_block.split('\n')
                option_start = None
                for i, line in enumerate(lines):
                    if re.match(r'^[A-H][\).:]\s', line.strip()):
                        option_start = i
                        break
                
                if option_start is None:
                    continue
                
                question_text = '\n'.join(lines[:option_start]).strip()
                question_text = re.sub(r'^\d+[\.\)]\s*', '', question_text)
                
                if not question_text or len(question_text) < 10:
                    continue
                
                options = {}
                for line in lines[option_start:]:
                    line = line.strip()
                    m = re.match(r'^([A-H])[\).:]\s*(.*)', line)
                    if m:
                        options[m.group(1)] = m.group(2).strip()
                
                if len(options) < 2:
                    continue
                
                qhash = get_question_text_hash(question_text)
                if qhash in existing_hashes:
                    continue
                
                qid = f"DABT-{start_id_num:04d}"
                start_id_num += 1
                
                cursor.execute(
                    "INSERT INTO questions (id, question_text, source_file_id) VALUES (?, ?, ?)",
                    (qid, question_text, source_file_id)
                )
                
                for letter in sorted(options.keys()):
                    cursor.execute(
                        "INSERT INTO answer_options (question_id, option_letter, option_text) VALUES (?, ?, ?)",
                        (qid, letter, options[letter])
                    )
                
                infer_topic_from_text(question_text, qid, cursor)
                existing_hashes[qhash] = qid
                docx_inserted += 1
                inserted += 1
                
                if inserted % 50 == 0:
                    conn.commit()
            
            conn.commit()
            log(f"  Inserted {docx_inserted} questions from {filename}")
    
    return start_id_num

def verify_database(conn):
    cursor = conn.cursor()
    stats = []
    
    cursor.execute("SELECT COUNT(*) FROM questions")
    total_q = cursor.fetchone()[0]
    stats.append(f"Total questions: {total_q}")
    
    stats.append("\n=== Distribution by Source ===")
    cursor.execute("""
        SELECT sf.bank_name, COUNT(*) as cnt
        FROM questions q
        JOIN source_files sf ON q.source_file_id = sf.id
        GROUP BY sf.id
        ORDER BY cnt DESC
    """)
    for row in cursor.fetchall():
        stats.append(f"  {row[0]}: {row[1]}")
    
    stats.append("\n=== Distribution by Domain ===")
    cursor.execute("""
        SELECT domain, COUNT(*) as cnt
        FROM question_domains
        GROUP BY domain
        ORDER BY cnt DESC
    """)
    for row in cursor.fetchall():
        stats.append(f"  {row[0]}: {row[1]}")
    
    cursor.execute("""
        SELECT COUNT(*) FROM questions q
        WHERE NOT EXISTS (SELECT 1 FROM question_domains d WHERE d.question_id = q.id)
    """)
    no_domain = cursor.fetchone()[0]
    stats.append(f"  (No domain assigned: {no_domain})")
    
    stats.append("\n=== Top 15 Topics ===")
    cursor.execute("""
        SELECT topic, COUNT(*) as cnt
        FROM question_topics
        GROUP BY topic
        ORDER BY cnt DESC
        LIMIT 15
    """)
    for row in cursor.fetchall():
        stats.append(f"  {row[0]}: {row[1]}")
    
    cursor.execute("SELECT COUNT(*) FROM answer_options")
    stats.append(f"\nTotal answer options: {cursor.fetchone()[0]}")
    
    cursor.execute("SELECT COUNT(*) FROM match_pairs")
    stats.append(f"Total match pairs: {cursor.fetchone()[0]}")
    
    cursor.execute("""
        SELECT COUNT(*) FROM questions q
        WHERE NOT EXISTS (SELECT 1 FROM answer_options a WHERE a.question_id = q.id)
    """)
    no_opts = cursor.fetchone()[0]
    stats.append(f"Questions with no options: {no_opts}")
    
    cursor.execute("""
        SELECT COUNT(*) FROM questions q
        WHERE q.correct_answer_letter IS NULL AND q.id NOT IN (
            SELECT question_id FROM match_pairs
        )
    """)
    no_ans = cursor.fetchone()[0]
    stats.append(f"Questions without answer: {no_ans}")
    
    stats.append("\n=== Spot Check (5 random) ===")
    cursor.execute("""
        SELECT q.id, q.question_text, q.correct_answer_letter, 
               (SELECT COUNT(*) FROM answer_options a WHERE a.question_id = q.id) as opt_count,
               sf.bank_name
        FROM questions q
        JOIN source_files sf ON q.source_file_id = sf.id
        ORDER BY RANDOM()
        LIMIT 5
    """)
    for row in cursor.fetchall():
        stats.append(f"\n  {row[0]} (from {row[4]})")
        qtext = (row[1][:100] + '...') if len(row[1]) > 100 else row[1]
        stats.append(f"    Q: {qtext}")
        stats.append(f"    Correct: {row[2]}")
        stats.append(f"    Options: {row[3]}")
        cursor.execute("SELECT option_letter, option_text FROM answer_options WHERE question_id = ? ORDER BY option_letter", (row[0],))
        for opt in cursor.fetchall():
            otext = (opt[1][:60] + '...') if len(opt[1]) > 60 else opt[1]
            stats.append(f"    {opt[0]}: {otext}")
    
    return '\n'.join(stats)

def main():
    log(f"=== DABT Database Build v2 - {datetime.now().isoformat()} ===")
    
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        log(f"Removed existing database at {DB_PATH}")
    
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    create_tables(conn)
    log("Tables created")
    
    classifications = load_classifications()
    
    # Source files
    cursor = conn.cursor()
    source_files = [
        (1, "Mini-ABT 1-11", "DABT_Practice_Questions_Database.xlsx", "mini-exam", "2024-2025", "Mini-ABT Practice Exams 1–11 (existing DB)"),
        (2, "2000Q Bank", "dabt_extract_2000q.csv", "comprehensive", "2024", "2000 Q practice bank"),
        (3, "Chapter Tests", "dabt_extract_chapter.csv", "chapter-based", "2024", "Chapter-based practice tests"),
        (4, "Kristen Mini Exams", "dabt_extract_mini.csv", "mini-exam", "2024", "Kristen's mini practice exams"),
        (5, "Kristen Topic Tests", "dabt_extract_topic.csv", "topic-based", "2024", "Kristen's topic-specific tests"),
        (6, "Past ABT Exams (2008-2014)", "2008-2014_Compiled_Recert_Exams.xlsx", "real-exam", "2008-2014", "Compiled recertification exams"),
        (7, "Past ABT Exams (PDFs)", "various PDFs", "real-exam", "2012-2017", "Additional past exam PDFs"),
    ]
    for sf in source_files:
        cursor.execute(
            "INSERT INTO source_files (id, bank_name, filename, format_type, year, description) VALUES (?, ?, ?, ?, ?, ?)", sf
        )
    conn.commit()
    log("Source files inserted")
    
    # Step 1: Load existing 446
    log("\n--- Step 1: Loading existing DB (Mini-ABT 1-11) ---")
    existing_hashes = load_existing_xlsx(conn, 1)
    
    # Step 2: 2000Q Bank
    log("\n--- Step 2: Loading 2000Q Bank ---")
    next_id = load_csv_data(CSV_2000Q, conn, 2, 447, existing_hashes)
    
    # Step 3: Chapter Tests
    log("\n--- Step 3: Loading Chapter Tests ---")
    next_id = load_csv_data(CSV_CHAPTER, conn, 3, next_id, existing_hashes)
    
    # Step 4: Kristen Mini Exams (dedup against existing Mini-ABT 1-11)
    log("\n--- Step 4: Loading Kristen Mini Exams (dedup) ---")
    # Re-use load_csv_data which does dedup via existing_hashes
    next_id = load_csv_data(CSV_MINI, conn, 4, next_id, existing_hashes)
    
    # Step 5: Kristen Topic Tests
    log("\n--- Step 5: Loading Kristen Topic Tests ---")
    next_id = load_csv_data(CSV_TOPIC, conn, 5, next_id, existing_hashes)
    
    # Step 6: Past ABT Exams (xlsx)
    log("\n--- Step 6: Loading Past ABT Exams (2008-2014) ---")
    next_id = parse_past_abt_exams(conn, 6, next_id, existing_hashes)
    
    # Step 7: PDF files (with dedup by file size)
    log("\n--- Step 7: Parsing PDF files ---")
    next_id = parse_pdf_files(conn, 7, next_id, existing_hashes)
    
    # Step 8: DOCX files
    log("\n--- Step 8: Parsing DOCX files ---")
    next_id = parse_docx_files(conn, 7, next_id, existing_hashes)
    
    # Step 9: Assign domains
    log("\n--- Step 9: Assigning domains ---")
    assign_domains(conn, classifications)
    
    # Step 10: Verify
    log("\n--- Step 10: Verification ---")
    stats_text = verify_database(conn)
    
    with open(STATS_PATH, 'w') as f:
        f.write(stats_text)
    log(f"\nStats written to {STATS_PATH}")
    
    with open(LOG_PATH, 'w') as f:
        f.write('\n'.join(log_lines))
    log(f"Log written to {LOG_PATH}")
    
    conn.close()
    log("\n=== DABT Database Build v2 Complete ===")

if __name__ == '__main__':
    main()
